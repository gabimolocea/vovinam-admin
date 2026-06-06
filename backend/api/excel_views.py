# api/excel_views.py
"""
Django REST Framework views for Excel import/export
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from .excel_sync import ExcelExportService, ExcelImportService, ExcelTemplateGenerator
from .models import Athlete
from .permissions import IsAdmin
import openpyxl


class ExcelSyncViewSet(viewsets.ViewSet):
    """
    Excel import/export endpoints
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download blank Excel template for athlete import.
        GET /api/excel/download_template/
        """
        wb = ExcelTemplateGenerator.create_athlete_template()
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="frvv_athletes_template.xlsx"'
        wb.save(response)
        
        return response
    
    @action(detail=False, methods=['get'])
    def export_athletes(self, request):
        """
        Export all athletes to Excel.
        GET /api/excel/export_athletes/?status=approved&club=1
        
        Query params:
        - status: Filter by status (approved, pending, etc.)
        - club: Filter by club ID
        - is_coach: Filter coaches (true/false)
        - is_referee: Filter referees (true/false)
        """
        queryset = Athlete.objects.all()
        
        # Apply filters
        if request.query_params.get('status'):
            queryset = queryset.filter(status=request.query_params['status'])
        
        if request.query_params.get('club'):
            queryset = queryset.filter(club_id=request.query_params['club'])
        
        if request.query_params.get('is_coach'):
            queryset = queryset.filter(is_coach=request.query_params['is_coach'].lower() == 'true')
        
        if request.query_params.get('is_referee'):
            queryset = queryset.filter(is_referee=request.query_params['is_referee'].lower() == 'true')
        
        # Generate filename
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'frvv_athletes_{timestamp}.xlsx'
        
        return ExcelExportService.export_to_http_response(queryset, filename)
    
    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def import_athletes(self, request):
        """
        Import athletes from Excel file.
        POST /api/excel/import_athletes/
        
        Body: multipart/form-data with 'file' field containing Excel file
        
        Response:
        {
            "success": true,
            "created": 10,
            "updated": 5,
            "errors": [...],
            "details": {...}
        }
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        
        # Validate file type
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'File must be an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Process import
        try:
            results = ExcelImportService.import_athletes(excel_file, request.user)
            
            return Response({
                'success': len(results['errors']) == 0,
                'created': len(results['created']),
                'updated': len(results['updated']),
                'errors': results['errors'],
                'skipped': len(results['skipped']),
                'details': {
                    'created_athletes': results['created'],
                    'updated_athletes': results['updated'],
                    'errors': results['errors']
                }
            })
        
        except Exception as e:
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def validate_import(self, request):
        """
        Validate Excel file without actually importing.
        POST /api/excel/validate_import/
        
        Returns validation errors and preview of changes.
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb['Athletes']
            
            from .excel_sync import AthleteExcelMapper
            
            validation_results = {
                'valid': [],
                'warnings': [],
                'errors': [],
                'summary': {
                    'total_rows': ws.max_row - 1,  # Exclude header
                    'new_athletes': 0,
                    'updates': 0,
                    'valid': 0,
                    'errors': 0
                }
            }
            
            # Get column mapping
            headers = {}
            for col_idx, (header_name, field_name) in enumerate(AthleteExcelMapper.COLUMNS, start=1):
                headers[col_idx] = field_name
            
            # Validate each row
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, field_name in headers.items():
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data[field_name] = cell_value
                
                # Skip empty rows
                if not row_data.get('first_name') and not row_data.get('last_name'):
                    continue
                
                # Convert and validate
                data, errors = AthleteExcelMapper.from_excel_row(row_data, row_idx)
                
                if errors:
                    validation_results['errors'].extend(errors)
                    validation_results['summary']['errors'] += 1
                else:
                    athlete_id = row_data.get('id')
                    if athlete_id:
                        # Check if exists
                        try:
                            athlete = Athlete.objects.get(pk=athlete_id)
                            validation_results['valid'].append({
                                'row': row_idx,
                                'action': 'update',
                                'name': f"{row_data['first_name']} {row_data['last_name']}"
                            })
                            validation_results['summary']['updates'] += 1
                        except Athlete.DoesNotExist:
                            validation_results['errors'].append(f"Row {row_idx}: Athlete ID {athlete_id} not found")
                            validation_results['summary']['errors'] += 1
                    else:
                        validation_results['valid'].append({
                            'row': row_idx,
                            'action': 'create',
                            'name': f"{row_data['first_name']} {row_data['last_name']}"
                        })
                        validation_results['summary']['new_athletes'] += 1
                    
                    validation_results['summary']['valid'] += 1
            
            validation_results['is_valid'] = len(validation_results['errors']) == 0
            
            return Response(validation_results)
        
        except Exception as e:
            return Response(
                {'error': f'Validation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def export_competitions(self, request):
        """
        Export competition results to Excel.
        GET /api/excel/export_competitions/?competition=1
        """
        from .models import CategoryAthleteScore
        from .excel_sync import CompetitionExcelMapper
        
        queryset = CategoryAthleteScore.objects.select_related(
            'category', 'category__competition', 'athlete', 'athlete__club'
        )
        
        if request.query_params.get('competition'):
            queryset = queryset.filter(category__competition_id=request.query_params['competition'])
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Competition Results'
        
        # Headers
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_idx, (header_name, _) in enumerate(CompetitionExcelMapper.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header_name
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row_idx, score in enumerate(queryset, start=2):
            row_data = CompetitionExcelMapper.to_excel_row(score)
            for col_idx, (_, field_name) in enumerate(CompetitionExcelMapper.COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx).value = row_data.get(field_name, '')
        
        # Save to response
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'frvv_competition_results_{timestamp}.xlsx'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
