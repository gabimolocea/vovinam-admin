# api/excel_sync.py
"""
Excel import/export functionality for bulk data management
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from .models import Athlete, Club, City, Grade, Category, GradeHistory
from io import BytesIO


class ExcelMapper:
    """Base class for Excel import/export mapping"""
    
    @staticmethod
    def to_excel_date(date_obj):
        """Convert Python date to Excel-friendly format"""
        if not date_obj:
            return None
        return date_obj.strftime('%Y-%m-%d') if hasattr(date_obj, 'strftime') else str(date_obj)
    
    @staticmethod
    def from_excel_date(date_str):
        """Convert Excel date string to Python date"""
        if not date_str:
            return None
        try:
            return datetime.strptime(str(date_str), '%Y-%m-%d').date()
        except:
            return None


class AthleteExcelMapper(ExcelMapper):
    """Map Athlete model to/from Excel"""
    
    COLUMNS = [
        ('ID', 'id'),
        ('First Name', 'first_name'),
        ('Last Name', 'last_name'),
        ('License Series', 'license_series'),
        ('Date of Birth', 'date_of_birth'),
        ('CNP', 'cnp'),
        ('Email', 'email'),
        ('Phone', 'mobile_number'),
        ('Club', 'club_name'),
        ('City', 'city_name'),
        ('Current Grade', 'current_grade_name'),
        ('Is Coach', 'is_coach'),
        ('Is Referee', 'is_referee'),
        ('Status', 'status'),
        ('Registered Date', 'registered_date'),
        ('Last Modified', 'updated_at'),
        ('Version', 'version'),
    ]
    
    @classmethod
    def to_excel_row(cls, athlete):
        """Convert Athlete instance to Excel row data"""
        return {
            'id': athlete.id,
            'first_name': athlete.first_name,
            'last_name': athlete.last_name,
            'license_series': athlete.license_series if hasattr(athlete, 'license_series') else '',
            'date_of_birth': cls.to_excel_date(athlete.date_of_birth),
            'cnp': athlete.cnp if hasattr(athlete, 'cnp') else '',
            'email': athlete.user.email if athlete.user else '',
            'mobile_number': athlete.mobile_number or '',
            'club_name': athlete.club.name if athlete.club else '',
            'city_name': athlete.city.name if athlete.city else '',
            'current_grade_name': athlete.current_grade.name if athlete.current_grade else '',
            'is_coach': 'Yes' if athlete.is_coach else 'No',
            'is_referee': 'Yes' if athlete.is_referee else 'No',
            'status': athlete.status,
            'registered_date': cls.to_excel_date(athlete.registered_date),
            'updated_at': cls.to_excel_date(getattr(athlete, 'updated_at', None) or athlete.submitted_date),
            'version': getattr(athlete, 'version', 1),
        }
    
    @classmethod
    def from_excel_row(cls, row_data, row_num=None):
        """
        Convert Excel row to Athlete data dict.
        Returns tuple: (data_dict, errors_list)
        """
        errors = []
        
        # Required fields
        if not row_data.get('first_name'):
            errors.append(f"Row {row_num}: First name is required")
        if not row_data.get('last_name'):
            errors.append(f"Row {row_num}: Last name is required")
        
        # Lookup foreign keys
        club = None
        if row_data.get('club_name'):
            club = Club.objects.filter(name=row_data['club_name']).first()
            if not club:
                errors.append(f"Row {row_num}: Club '{row_data['club_name']}' not found")
        
        city = None
        if row_data.get('city_name'):
            city = City.objects.filter(name=row_data['city_name']).first()
            if not city:
                errors.append(f"Row {row_num}: City '{row_data['city_name']}' not found")
        
        grade = None
        if row_data.get('current_grade_name'):
            grade = Grade.objects.filter(name=row_data['current_grade_name']).first()
            if not grade:
                errors.append(f"Row {row_num}: Grade '{row_data['current_grade_name']}' not found")
        
        data = {
            'first_name': row_data.get('first_name', ''),
            'last_name': row_data.get('last_name', ''),
            'license_series': row_data.get('license_series', ''),
            'date_of_birth': cls.from_excel_date(row_data.get('date_of_birth')),
            'cnp': row_data.get('cnp', ''),
            'mobile_number': row_data.get('mobile_number', ''),
            'club': club.id if club else None,
            'city': city.id if city else None,
            'current_grade': grade.id if grade else None,
            'is_coach': row_data.get('is_coach', '').lower() in ['yes', 'true', '1'],
            'is_referee': row_data.get('is_referee', '').lower() in ['yes', 'true', '1'],
            'status': row_data.get('status', 'approved'),
            'registered_date': cls.from_excel_date(row_data.get('registered_date')),
        }
        
        return data, errors


class CompetitionExcelMapper(ExcelMapper):
    """Map Competition results to/from Excel"""
    
    COLUMNS = [
        ('ID', 'id'),
        ('Competition', 'competition_name'),
        ('Category', 'category_name'),
        ('Athlete', 'athlete_name'),
        ('Club', 'club_name'),
        ('Placement', 'placement'),
        ('Score', 'score'),
        ('Status', 'status'),
        ('Date', 'competition_date'),
        ('Last Modified', 'updated_at'),
    ]
    
    @classmethod
    def to_excel_row(cls, score):
        """Convert CategoryAthleteScore to Excel row"""
        from .models import CategoryAthleteScore
        return {
            'id': score.id,
            'competition_name': score.category.competition.name if score.category and score.category.competition else '',
            'category_name': score.category.name if score.category else '',
            'athlete_name': f"{score.athlete.first_name} {score.athlete.last_name}" if score.athlete else '',
            'club_name': score.athlete.club.name if score.athlete and score.athlete.club else '',
            'placement': score.placement_claimed,
            'score': str(score.final_score),
            'status': score.status,
            'competition_date': cls.to_excel_date(score.category.competition.event.start_date if score.category and score.category.competition and score.category.competition.event else None),
            'updated_at': cls.to_excel_date(score.submitted_date),
        }


class ExcelTemplateGenerator:
    """Generate Excel templates with validation and formatting"""
    
    @staticmethod
    def create_athlete_template():
        """Create Excel template for athlete bulk import"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Athletes'
        
        # Headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        columns = AthleteExcelMapper.COLUMNS
        for col_idx, (header_name, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        widths = [8, 15, 15, 12, 15, 25, 15, 20, 15, 15, 10, 12, 12, 15, 20, 8]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        
        # Add data validation for dropdowns
        status_validator = DataValidation(
            type='list',
            formula1='"pending,approved,rejected,revision_required"',
            allow_blank=True
        )
        ws.add_data_validation(status_validator)
        status_validator.add(f'M2:M1000')  # Status column
        
        bool_validator = DataValidation(
            type='list',
            formula1='"Yes,No"',
            allow_blank=True
        )
        ws.add_data_validation(bool_validator)
        bool_validator.add(f'K2:K1000')  # Is Coach
        bool_validator.add(f'L2:L1000')  # Is Referee
        
        # Add reference data sheets
        ExcelTemplateGenerator._add_clubs_sheet(wb)
        ExcelTemplateGenerator._add_cities_sheet(wb)
        ExcelTemplateGenerator._add_grades_sheet(wb)
        ExcelTemplateGenerator._add_instructions_sheet(wb)
        
        return wb
    
    @staticmethod
    def _add_clubs_sheet(workbook):
        """Add sheet with available clubs for reference"""
        ws = workbook.create_sheet('Clubs Reference')
        ws['A1'] = 'Club Name'
        ws['A1'].font = Font(bold=True)
        
        for idx, club in enumerate(Club.objects.all().order_by('name'), start=2):
            ws[f'A{idx}'] = club.name
    
    @staticmethod
    def _add_cities_sheet(workbook):
        """Add sheet with available cities for reference"""
        ws = workbook.create_sheet('Cities Reference')
        ws['A1'] = 'City Name'
        ws['A1'].font = Font(bold=True)
        
        for idx, city in enumerate(City.objects.all().order_by('name'), start=2):
            ws[f'A{idx}'] = city.name
    
    @staticmethod
    def _add_grades_sheet(workbook):
        """Add sheet with available grades for reference"""
        ws = workbook.create_sheet('Grades Reference')
        ws['A1'] = 'Grade Name'
        ws['B1'] = 'Rank Order'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        for idx, grade in enumerate(Grade.objects.all().order_by('rank_order'), start=2):
            ws[f'A{idx}'] = grade.name
            ws[f'B{idx}'] = grade.rank_order
    
    @staticmethod
    def _add_instructions_sheet(workbook):
        """Add instructions sheet"""
        ws = workbook.create_sheet('Instructions')
        
        instructions = [
            ('FRVV Athletes Excel Template', 'title'),
            ('', 'blank'),
            ('How to use this template:', 'heading'),
            ('1. Download latest athlete data using the sync button in the app', 'text'),
            ('2. Edit athlete data in the Athletes sheet (do NOT modify ID, Version, or Last Modified columns)', 'text'),
            ('3. Add new athletes by leaving the ID column blank', 'text'),
            ('4. Use the reference sheets (Clubs, Cities, Grades) to ensure correct spelling', 'text'),
            ('5. Upload the file back to the app to sync changes', 'text'),
            ('', 'blank'),
            ('Important Notes:', 'heading'),
            ('- ID column: Leave blank for new athletes, do not modify for existing athletes', 'text'),
            ('- Version column: Used for conflict detection, do not modify', 'text'),
            ('- Status: Use dropdown (pending, approved, rejected, revision_required)', 'text'),
            ('- Is Coach/Is Referee: Use dropdown (Yes/No)', 'text'),
            ('- Dates: Use format YYYY-MM-DD (e.g., 2024-01-15)', 'text'),
            ('', 'blank'),
            ('Conflict Resolution:', 'heading'),
            ('If the server data changed since you downloaded, you will be notified of conflicts.', 'text'),
            ('You can choose to: Keep server data, Use your changes, or Merge manually', 'text'),
        ]
        
        for row_idx, (text, style) in enumerate(instructions, start=1):
            cell = ws[f'A{row_idx}']
            cell.value = text
            
            if style == 'title':
                cell.font = Font(size=16, bold=True, color='366092')
            elif style == 'heading':
                cell.font = Font(size=12, bold=True)
            elif style == 'text':
                cell.alignment = Alignment(wrap_text=True)
        
        ws.column_dimensions['A'].width = 80


class ExcelImportService:
    """Service for importing Excel files"""
    
    @staticmethod
    def import_athletes(file_obj, user=None):
        """
        Import athletes from Excel file.
        Returns dict with success/error counts and details.
        """
        wb = load_workbook(file_obj)
        ws = wb['Athletes']
        
        results = {
            'created': [],
            'updated': [],
            'errors': [],
            'skipped': []
        }
        
        # Get column mapping from header row
        headers = {}
        for col_idx, (header_name, field_name) in enumerate(AthleteExcelMapper.COLUMNS, start=1):
            headers[col_idx] = field_name
        
        # Process each row
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, field_name in headers.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[field_name] = cell_value
            
            # Skip empty rows
            if not row_data.get('first_name') and not row_data.get('last_name'):
                continue
            
            # Convert to athlete data
            data, errors = AthleteExcelMapper.from_excel_row(row_data, row_idx)
            
            if errors:
                results['errors'].extend(errors)
                continue
            
            # Create or update
            athlete_id = row_data.get('id')
            
            try:
                if athlete_id:
                    # Update existing
                    athlete = Athlete.objects.get(pk=athlete_id)
                    
                    # Check version for conflict
                    excel_version = row_data.get('version', 1)
                    if hasattr(athlete, 'version') and athlete.version != excel_version:
                        results['errors'].append(
                            f"Row {row_idx}: Conflict detected. Server version {athlete.version}, Excel version {excel_version}"
                        )
                        continue
                    
                    # Update fields
                    for key, value in data.items():
                        if value is not None:
                            setattr(athlete, key, value)
                    
                    if user and hasattr(athlete, 'excel_row_number'):
                        athlete.mark_excel_import(row_idx, user)
                    
                    athlete.save()
                    results['updated'].append({
                        'id': athlete.id,
                        'name': f"{athlete.first_name} {athlete.last_name}",
                        'row': row_idx
                    })
                else:
                    # Create new
                    athlete = Athlete.objects.create(**data)
                    
                    if user and hasattr(athlete, 'excel_row_number'):
                        athlete.mark_excel_import(row_idx, user)
                    
                    results['created'].append({
                        'id': athlete.id,
                        'name': f"{athlete.first_name} {athlete.last_name}",
                        'row': row_idx
                    })
            
            except Athlete.DoesNotExist:
                results['errors'].append(f"Row {row_idx}: Athlete ID {athlete_id} not found")
            except Exception as e:
                results['errors'].append(f"Row {row_idx}: {str(e)}")
        
        return results


class ExcelExportService:
    """Service for exporting data to Excel"""
    
    @staticmethod
    def export_athletes(queryset=None):
        """Export athletes to Excel file"""
        if queryset is None:
            queryset = Athlete.objects.all()
        
        # Create workbook from template
        wb = ExcelTemplateGenerator.create_athlete_template()
        ws = wb['Athletes']
        
        # Add data
        for row_idx, athlete in enumerate(queryset.select_related('club', 'city', 'current_grade', 'user'), start=2):
            row_data = AthleteExcelMapper.to_excel_row(athlete)
            
            for col_idx, (_, field_name) in enumerate(AthleteExcelMapper.COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx).value = row_data.get(field_name, '')
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_to_http_response(queryset=None, filename='athletes_export.xlsx'):
        """Export to HTTP response for download"""
        output = ExcelExportService.export_athletes(queryset)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
