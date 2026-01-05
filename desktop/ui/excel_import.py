"""
Excel import dialog
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QFileDialog, QTextEdit, QMessageBox
)
from PyQt6.QtCore import Qt
from openpyxl import load_workbook
from models.db import Database
from datetime import datetime
import config

class ExcelImportDialog(QDialog):
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.db = db
        self.file_path = None
        self.init_ui()
    
    def init_ui(self):
        """Initialize the dialog"""
        self.setWindowTitle('Import Athletes from Excel')
        self.setMinimumSize(600, 400)
        
        layout = QVBoxLayout(self)
        
        # File selection
        file_layout = QHBoxLayout()
        self.file_label = QLabel('No file selected')
        file_layout.addWidget(self.file_label)
        
        btn_browse = QPushButton('Browse...')
        btn_browse.clicked.connect(self.browse_file)
        file_layout.addWidget(btn_browse)
        
        layout.addLayout(file_layout)
        
        # Preview area
        layout.addWidget(QLabel('Preview:'))
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        layout.addWidget(self.preview_text)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        btn_validate = QPushButton('Validate')
        btn_validate.clicked.connect(self.validate_file)
        button_layout.addWidget(btn_validate)
        
        btn_import = QPushButton('Import')
        btn_import.clicked.connect(self.import_file)
        button_layout.addWidget(btn_import)
        
        btn_cancel = QPushButton('Cancel')
        btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(btn_cancel)
        
        layout.addLayout(button_layout)
    
    def browse_file(self):
        """Browse for Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 'Select Excel File',
            str(config.EXCEL_DIR),
            'Excel Files (*.xlsx *.xls)'
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(file_path)
            self.preview_file()
    
    def preview_file(self):
        """Preview Excel file contents"""
        if not self.file_path:
            return
        
        try:
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            preview = f"Sheet: {ws.title}\n"
            preview += f"Rows: {ws.max_row}\n\n"
            
            # Show first 5 rows
            for row_idx, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
                preview += f"Row {row_idx}: {row}\n"
            
            self.preview_text.setText(preview)
            wb.close()
            
        except Exception as e:
            self.preview_text.setText(f"Error reading file: {str(e)}")
    
    def validate_file(self):
        """Validate Excel file without importing"""
        if not self.file_path:
            QMessageBox.warning(self, 'No File', 'Please select a file first')
            return
        
        try:
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            
            errors = []
            valid_count = 0
            
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                first_name = row[0] if len(row) > 0 else None
                last_name = row[1] if len(row) > 1 else None
                
                if not first_name or not last_name:
                    errors.append(f"Row {row_idx}: Missing name")
                else:
                    valid_count += 1
            
            wb.close()
            
            result = f"Validation Results:\n\n"
            result += f"Valid rows: {valid_count}\n"
            result += f"Errors: {len(errors)}\n\n"
            
            if errors:
                result += "Errors:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    result += f"\n... and {len(errors) - 10} more"
            
            QMessageBox.information(self, 'Validation Complete', result)
            
        except Exception as e:
            QMessageBox.critical(self, 'Validation Error', f'Error: {str(e)}')
    
    def import_file(self):
        """Import athletes from Excel file"""
        if not self.file_path:
            QMessageBox.warning(self, 'No File', 'Please select a file first')
            return
        
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            imported = 0
            errors = []
            
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    first_name = row[0] if len(row) > 0 else None
                    last_name = row[1] if len(row) > 1 else None
                    
                    if not first_name or not last_name:
                        continue
                    
                    athlete_data = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'date_of_birth': row[2] if len(row) > 2 else None,
                        'mobile_number': row[3] if len(row) > 3 else None,
                        'club_name': row[4] if len(row) > 4 else None,
                        'city_name': row[5] if len(row) > 5 else None,
                        'status': row[6] if len(row) > 6 else 'pending',
                        'temp_id': f"temp_{datetime.now().timestamp()}_{imported}",
                        'created_offline': 1,
                        'is_synced': 0
                    }
                    
                    self.db.insert_athlete(athlete_data)
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            wb.close()
            
            result = f"Import Complete!\n\n"
            result += f"Imported: {imported} athletes\n"
            result += f"Errors: {len(errors)}"
            
            QMessageBox.information(self, 'Import Complete', result)
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, 'Import Error', f'Error: {str(e)}')
