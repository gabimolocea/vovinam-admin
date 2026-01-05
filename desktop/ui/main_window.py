"""
Main application window
"""
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QStatusBar, QMenuBar, QMenu,
    QMessageBox, QInputDialog, QFileDialog, QApplication
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QAction
import config
from models.db import Database
from sync.sync_manager import SyncManager
from ui.athlete_list import AthleteListWidget
from ui.excel_import import ExcelImportDialog
from ui.login_dialog import LoginDialog
from ui.competitions_tab import CompetitionsTab
from ui.categories_tab import CategoriesTab
from ui.matches_tab import MatchesTab
from ui.clubs_tab import ClubsTab
from ui.grades_tab import GradesTab
from ui.theme import ThemeManager

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.sync_manager = SyncManager()
        self.is_logged_in = False
        self.current_user = None
        self.init_ui()
        self.load_data()
        
        # Prompt for login on startup
        self.prompt_login()
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle(config.WINDOW_TITLE)
        self.setGeometry(100, 100, config.WINDOW_WIDTH, config.WINDOW_HEIGHT)
        
        # Create menu bar
        self.create_menus()
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        layout = QVBoxLayout(central_widget)
        
        # Toolbar
        toolbar = self.create_toolbar()
        layout.addLayout(toolbar)
        
        # Tab widget for different sections
        from PyQt6.QtWidgets import QTabWidget
        self.tabs = QTabWidget()
        
        # Dashboard tab (first)
        from ui.dashboard import DashboardWidget
        self.dashboard = DashboardWidget(self.db)
        self.tabs.addTab(self.dashboard, '📊 Dashboard')
        
        # Athletes tab
        self.athlete_list = AthleteListWidget(self.db)
        self.tabs.addTab(self.athlete_list, '👥 Athletes')
        
        # Competitions tab
        self.competitions_tab = CompetitionsTab(self.db, self.sync_manager)
        self.tabs.addTab(self.competitions_tab, '🏆 Competitions')
        
        # Categories tab
        self.categories_tab = CategoriesTab(self.db)
        self.tabs.addTab(self.categories_tab, '📋 Categories')
        
        # Matches tab
        self.matches_tab = MatchesTab(self.db)
        # Clubs tab
        self.clubs_tab = ClubsTab(self.db)
        self.tabs.addTab(self.clubs_tab, '🏢 Clubs')
        
        # Grades tab
        self.grades_tab = GradesTab(self.db)
        self.tabs.addTab(self.grades_tab, '🥋 Grades')
        
        self.tabs.addTab(self.matches_tab, '🥊 Matches')
        
        layout.addWidget(self.tabs)
        
        # Status bar
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.update_status()
    
    def create_menus(self):
        """Create menu bar"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu('&File')
        
        import_action = QAction('&Import Excel...', self)
        import_action.triggered.connect(self.import_excel)
        file_menu.addAction(import_action)
        
        export_action = QAction('&Export Excel...', self)
        export_action.triggered.connect(self.export_excel)
        file_menu.addAction(export_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction('E&xit', self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Sync menu
        sync_menu = menubar.addMenu('&Sync')
        
        login_action = QAction('&Login...', self)
        login_action.triggered.connect(self.prompt_login)
        sync_menu.addAction(login_action)
        
        logout_action = QAction('Log&out', self)
        logout_action.triggered.connect(self.logout)
        sync_menu.addAction(logout_action)
        
        sync_menu.addSeparator()
        
        test_conn_action = QAction('Test &Connection', self)
        test_conn_action.triggered.connect(self.test_connection)
        sync_menu.addAction(test_conn_action)
        
        push_action = QAction('&Upload to Server', self)
        push_action.triggered.connect(self.push_data)
        sync_menu.addAction(push_action)
        
        sync_all_action = QAction('&Full Sync (Both Ways)', self)
        sync_all_action.triggered.connect(self.full_sync)
        sync_menu.addAction(sync_all_action)
        
        # View menu
        view_menu = menubar.addMenu('&View')
        
        light_theme_action = QAction('☀️ &Light Mode', self)
        light_theme_action.triggered.connect(lambda: self.change_theme('light'))
        view_menu.addAction(light_theme_action)
        
        dark_theme_action = QAction('🌙 &Dark Mode', self)
        dark_theme_action.triggered.connect(lambda: self.change_theme('dark'))
        view_menu.addAction(dark_theme_action)
        
        # Help menu
        help_menu = menubar.addMenu('&Help')
        
        about_action = QAction('&About', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def create_toolbar(self) -> QHBoxLayout:
        """Create toolbar with action buttons"""
        toolbar = QHBoxLayout()
        
        # Sync buttons
        btn_sync = QPushButton('🔄 Sync All')
        btn_sync.clicked.connect(self.full_sync)
        btn_sync.setToolTip('Download all data from server and upload local changes')
        toolbar.addWidget(btn_sync)
        
        btn_push = QPushButton('⬆ Upload Only')
        btn_push.clicked.connect(self.push_data)
        btn_push.setToolTip('Upload local changes to server without downloading')
        toolbar.addWidget(btn_push)
        
        toolbar.addStretch()
        
        # Excel buttons
        btn_import = QPushButton('📥 Import Excel')
        btn_import.clicked.connect(self.import_excel)
        toolbar.addWidget(btn_import)
        
        btn_export = QPushButton('📤 Export Excel')
        btn_export.clicked.connect(self.export_excel)
        toolbar.addWidget(btn_export)
        
        toolbar.addStretch()
        
        # Refresh button
        btn_refresh = QPushButton('🔄 Refresh')
        btn_refresh.clicked.connect(self.load_data)
        toolbar.addWidget(btn_refresh)
        
        return toolbar
    
    def load_data(self):
        """Load all data from database"""
        self.dashboard.load_data()
        self.athlete_list.load_athletes()
        self.competitions_tab.load_competitions()
        self.categories_tab.load_categories()
        self.matches_tab.load_matches()
        self.clubs_tab.load_clubs()
        self.grades_tab.load_grades()
        self.update_status()
    
    def update_status(self):
        """Update status bar"""
        athletes = self.db.get_all_athletes()
        unsynced = self.db.get_unsynced_athletes()
        
        login_status = f"👤 {self.current_user.get('email', 'Not logged in')}" if self.is_logged_in else "🔒 Not logged in"
        status_text = f"{login_status} | Athletes: {len(athletes)} | Unsynced: {len(unsynced)}"
        self.statusBar.showMessage(status_text)
    
    def prompt_login(self):
        """Show login dialog"""
        dialog = LoginDialog(self)
        if dialog.exec():
            token = dialog.get_token()
            self.current_user = dialog.get_user_info()
            
            if token:
                self.sync_manager.set_auth_token(token)
                self.is_logged_in = True
                self.update_status()
                QMessageBox.information(
                    self, 'Login Success',
                    f"Logged in as {self.current_user.get('email', 'User')}"
                )
    
    def logout(self):
        """Logout from API"""
        self.sync_manager.set_auth_token(None)
        self.is_logged_in = False
        self.current_user = None
        self.update_status()
        QMessageBox.information(self, 'Logged Out', 'You have been logged out')
    
    def check_login(self) -> bool:
        """Check if user is logged in"""
        if not self.is_logged_in:
            reply = QMessageBox.question(
                self, 'Login Required',
                'You need to login to sync with the server.\n\nLogin now?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.prompt_login()
                return self.is_logged_in
            return False
        return True
    
    def load_data(self):
        """Load all data from database"""
        self.athlete_list.load_athletes()
        self.competitions_tab.load_competitions()
        self.categories_tab.load_categories()
        self.matches_tab.load_matches()
        self.clubs_tab.load_clubs()
        self.grades_tab.load_grades()
        self.update_status()
    
    def update_status(self):
        """Update status bar"""
        athletes = self.db.get_all_athletes()
        unsynced = self.db.get_unsynced_athletes()
        
        status_text = f"Athletes: {len(athletes)} | Unsynced: {len(unsynced)}"
        self.statusBar.showMessage(status_text)
    
    def test_connection(self):
        """Test connection to API"""
        success, message = self.sync_manager.test_connection()
        
        if success:
            QMessageBox.information(self, 'Connection Test', f'✓ {message}')
        else:
            QMessageBox.warning(self, 'Connection Test', f'✗ {message}')
    
    def pull_data(self):
        """Download data from server"""
        if not self.check_login():
            return
        
        self.statusBar.showMessage('Downloading from server...')
        
        # First sync reference data (cities, clubs, grades)
        success, msg = self.sync_manager.sync_reference_data()
        if not success:
            QMessageBox.warning(self, 'Sync Error', f'Failed to sync reference data: {msg}')
            return
        
        print(f"Reference data sync: {msg}")
        
        # Pull all data types
        total_count = 0
        messages = [msg]  # Include reference data message
        
        # Athletes
        success, msg, count = self.sync_manager.pull_athletes()
        if success:
            total_count += count
            messages.append(f"Athletes: {count}")
        
        # Competitions
        success, msg, count = self.sync_manager.pull_competitions()
        if success:
            total_count += count
            messages.append(f"Competitions: {count}")
        
        # Categories
        success, msg, count = self.sync_manager.pull_categories()
        if success:
            total_count += count
            messages.append(f"Categories: {count}")
        
        # Matches
        success, msg, count = self.sync_manager.pull_matches()
        if success:
            total_count += count
            messages.append(f"Matches: {count}")
        
        # Grade History
        success, msg, count = self.sync_manager.pull_grade_history()
        if success:
            total_count += count
            messages.append(f"Grade History: {count}")
        
        # Visas
        success, msg, count = self.sync_manager.pull_visas()
        if success:
            total_count += count
            messages.append(f"Visas: {count}")
        else:
            messages.append(f"Visas: ERROR - {msg}")
        
        # Athlete Results
        success, msg, count = self.sync_manager.pull_athlete_results()
        if success:
            total_count += count
            messages.append(f"Results: {count}")
        
        QMessageBox.information(
            self, 'Download Complete',
            f"Downloaded {total_count} records:\n" + "\n".join(messages)
        )
        self.load_data()
        
        self.statusBar.showMessage('Ready')
    
    def push_data(self):
        """Upload all local changes to server"""
        if not self.check_login():
            return
        
        # Count unsynced items
        unsynced_competitions = self.db.get_unsynced_competitions()
        unsynced_categories = self.db.get_unsynced_categories()
        unsynced_grades = self.db.get_unsynced_grades()
        unsynced_clubs = self.db.get_unsynced_clubs()
        unsynced_athletes = self.db.get_unsynced_athletes()
        unsynced_matches = self.db.get_unsynced_matches()
        
        conn = self.db.connect()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM referee_scores 
            WHERE server_id IS NULL
        """)
        unsynced_scores = cursor.fetchone()[0]
        
        # Get pending deletions
        pending_deletions = self.db.get_pending_deletions()
        
        total_unsynced = len(unsynced_competitions) + len(unsynced_categories) + len(unsynced_grades) + len(unsynced_clubs) + len(unsynced_athletes) + len(unsynced_matches) + unsynced_scores + len(pending_deletions)
        
        if total_unsynced == 0:
            QMessageBox.information(self, 'Upload', 'No changes to upload')
            return
        
        # Show upload dialog with checkboxes
        from ui.upload_dialog import UploadDialog
        dialog = UploadDialog(
            self,
            counts={
                'competitions': len(unsynced_competitions),
                'categories': len(unsynced_categories),
                'grades': len(unsynced_grades),
                'clubs': len(unsynced_clubs),
                'athletes': len(unsynced_athletes),
                'matches': len(unsynced_matches),
                'scores': unsynced_scores,
                'deletions': len(pending_deletions)
            }
        )
        
        if dialog.exec() != dialog.DialogCode.Accepted:
            return
        
        selected = dialog.get_selected()
        
        messages = []
        total_uploaded = 0
        
        # Upload competitions
        if selected['competitions'] and unsynced_competitions:
            self.statusBar.showMessage('Uploading competitions...')
            success, msg, count = self.sync_manager.push_competitions()
            if success:
                total_uploaded += count
                messages.append(f"Competitions: {count}")
            else:
                messages.append(f"Competitions: {msg}")
        
        # Upload categories
        if selected['categories'] and unsynced_categories:
            self.statusBar.showMessage('Uploading categories...')
            success, msg, count = self.sync_manager.push_categories()
            if success:
                total_uploaded += count
                messages.append(f"Categories: {count}")
            else:
                messages.append(f"Categories: {msg}")
        
        # Upload grades
        if selected['grades'] and unsynced_grades:
            self.statusBar.showMessage('Uploading grades...')
            success, msg, count = self.sync_manager.push_grades()
            if success:
                total_uploaded += count
                messages.append(f"Grades: {count}")
            else:
                messages.append(f"Grades: {msg}")
        
        # Upload clubs
        if selected['clubs'] and unsynced_clubs:
            self.statusBar.showMessage('Uploading clubs...')
            success, msg, count = self.sync_manager.push_clubs()
            if success:
                total_uploaded += count
                messages.append(f"Clubs: {count}")
            else:
                messages.append(f"Clubs: {msg}")
        
        # Upload athletes
        if selected['athletes'] and unsynced_athletes:
            self.statusBar.showMessage('Uploading athletes...')
            success, msg, count = self.sync_manager.push_athletes()
            if success:
                total_uploaded += count
                messages.append(f"Athletes: {count}")
        
        # Upload matches (central referee changes)
        if selected['matches'] and unsynced_matches:
            self.statusBar.showMessage('Uploading match changes...')
            success, msg, count = self.sync_manager.push_matches()
            if success:
                total_uploaded += count
                messages.append(f"Matches: {count}")
            else:
                messages.append(f"Matches: {msg}")
        
        # Upload referee scores
        if selected['scores'] and unsynced_scores > 0:
            self.statusBar.showMessage('Uploading referee scores...')
            success, msg, count = self.sync_manager.push_referee_scores()
            if success:
                total_uploaded += count
                messages.append(f"Referee Scores: {count}")
        
        # Upload deletions
        if selected['deletions'] and pending_deletions:
            self.statusBar.showMessage('Syncing deletions...')
            success, msg, count = self.sync_manager.push_deletions()
            if success:
                total_uploaded += count
                messages.append(f"Deletions: {count}")
            else:
                messages.append(f"Deletions: {msg}")
        
        if messages:
            result = f"Upload Complete!\n\nUploaded {total_uploaded} records:\n" + "\n".join(messages)
            QMessageBox.information(self, 'Upload Complete', result)
        else:
            QMessageBox.information(self, 'Upload', 'No items selected to upload')
        
        self.load_data()
        
        self.statusBar.showMessage('Ready')
    
    def full_sync(self):
        """Perform full bidirectional sync for all entities"""
        if not self.check_login():
            return
        
        self.statusBar.showMessage('Full sync in progress...')
        
        messages = []
        total_pulled = 0
        total_pushed = 0
        
        # Step 1: Pull all data from server
        self.statusBar.showMessage('Downloading reference data...')
        success, msg = self.sync_manager.sync_reference_data()
        if success:
            messages.append(msg)
        
        # Pull athletes
        self.statusBar.showMessage('Downloading athletes...')
        success, msg, count = self.sync_manager.pull_athletes()
        if success:
            total_pulled += count
            messages.append(f"Downloaded {count} athletes")
        
        # Pull competitions
        self.statusBar.showMessage('Downloading competitions...')
        success, msg, count = self.sync_manager.pull_competitions()
        if success:
            total_pulled += count
            messages.append(f"Downloaded {count} competitions")
        
        # Pull categories
        self.statusBar.showMessage('Downloading categories...')
        success, msg, count = self.sync_manager.pull_categories()
        if success:
            total_pulled += count
            messages.append(f"Downloaded {count} categories")
        
        # Pull matches
        self.statusBar.showMessage('Downloading matches...')
        success, msg, count = self.sync_manager.pull_matches()
        if success:
            total_pulled += count
            messages.append(f"Downloaded {count} matches")
        
        # Step 2: Push local changes to server
        self.statusBar.showMessage('Uploading competitions...')
        success, msg, count = self.sync_manager.push_competitions()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Uploaded {count} competitions")
        
        self.statusBar.showMessage('Uploading categories...')
        success, msg, count = self.sync_manager.push_categories()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Uploaded {count} categories")
        
        self.statusBar.showMessage('Uploading grades...')
        success, msg, count = self.sync_manager.push_grades()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Uploaded {count} grades")
        
        self.statusBar.showMessage('Uploading clubs...')
        success, msg, count = self.sync_manager.push_clubs()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Uploaded {count} clubs")
        
        self.statusBar.showMessage('Uploading athletes...')
        success, msg, count = self.sync_manager.push_athletes()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Uploaded {count} athletes")
        
        # Push referee scores
        self.statusBar.showMessage('Uploading referee scores...')
        success, msg, count = self.sync_manager.push_referee_scores()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Uploaded {count} referee scores")
        
        # Push deletions
        self.statusBar.showMessage('Syncing deletions...')
        success, msg, count = self.sync_manager.push_deletions()
        if success and count > 0:
            total_pushed += count
            messages.append(f"Synced {count} deletions")
        
        # Show results
        result = f"Full Sync Complete!\n\n"
        result += f"Total downloaded: {total_pulled}\n"
        result += f"Total uploaded: {total_pushed}\n\n"
        result += "\n".join(messages)
        QMessageBox.information(self, 'Sync Complete', result)
        self.load_data()
        
        self.statusBar.showMessage('Ready')
    
    def import_excel(self):
        """Import data from Excel - intelligently detect entity type from sheet names"""
        from openpyxl import load_workbook
        from datetime import datetime
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, 'Import Excel File',
            str(config.EXCEL_DIR),
            'Excel Files (*.xlsx *.xls)'
        )
        
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path)
            sheet_names = wb.sheetnames
            
            # Let user choose which sheet to import
            from PyQt6.QtWidgets import QInputDialog
            sheet_name, ok = QInputDialog.getItem(
                self, 'Select Sheet', 
                'Choose which data to import:',
                sheet_names, 0, False
            )
            
            if not ok:
                return
            
            ws = wb[sheet_name]
            
            # Detect entity type and import accordingly
            sheet_lower = sheet_name.lower()
            
            if 'athlete' in sheet_lower:
                self._import_athletes_sheet(ws)
            elif 'competition' in sheet_lower or 'event' in sheet_lower:
                self._import_competitions_sheet(ws)
            elif 'categor' in sheet_lower:
                self._import_categories_sheet(ws)
            elif 'match' in sheet_lower:
                self._import_matches_sheet(ws)
            elif 'club' in sheet_lower:
                self._import_clubs_sheet(ws)
            elif 'grade' in sheet_lower:
                self._import_grades_sheet(ws)
            else:
                # Default to athletes
                self._import_athletes_sheet(ws)
            
            wb.close()
            self.load_data()
            
        except Exception as e:
            QMessageBox.critical(self, 'Import Error', f'Error: {str(e)}')
    
    def _import_athletes_sheet(self, ws):
        """Import athletes from worksheet"""
        from datetime import datetime
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row[0] or not row[1]:  # Skip if no name
                    continue
                
                athlete_data = {
                    'first_name': str(row[0]) if row[0] else None,
                    'last_name': str(row[1]) if row[1] else None,
                    'date_of_birth': str(row[2]) if len(row) > 2 and row[2] else None,
                    'mobile_number': str(row[3]) if len(row) > 3 and row[3] else None,
                    'club_name': str(row[4]) if len(row) > 4 and row[4] else None,
                    'city_name': str(row[5]) if len(row) > 5 and row[5] else None,
                    'current_grade_name': str(row[6]) if len(row) > 6 and row[6] else None,
                    'status': str(row[7]) if len(row) > 7 and row[7] else 'pending',
                    'temp_id': f"temp_{datetime.now().timestamp()}_{imported}",
                    'created_offline': 1,
                    'is_synced': 0
                }
                
                self.db.insert_athlete(athlete_data)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}"[:100])
        
        result = f"Athletes Import Complete!\n\nImported: {imported}\nErrors: {len(errors)}"
        if errors:
            result += "\n\nFirst errors:\n" + "\n".join(errors[:5])
        
        QMessageBox.information(self, 'Import Complete', result)
    
    def _import_competitions_sheet(self, ws):
        """Import competitions - read-only message"""
        QMessageBox.information(
            self, 'Import Not Available',
            'Competitions are managed on the server.\n'
            'Use the web application to create/edit competitions.'
        )
    
    def _import_categories_sheet(self, ws):
        """Import categories - read-only message"""
        QMessageBox.information(
            self, 'Import Not Available',
            'Categories are managed on the server.\n'
            'Use the web application to create/edit categories.'
        )
    
    def _import_matches_sheet(self, ws):
        """Import matches - read-only message"""
        QMessageBox.information(
            self, 'Import Not Available',
            'Matches are managed on the server.\n'
            'Use the web application to create/edit matches.'
        )
    
    def _import_clubs_sheet(self, ws):
        """Import clubs - read-only message"""
        QMessageBox.information(
            self, 'Import Not Available',
            'Clubs are managed on the server.\n'
            'Use the web application to create/edit clubs.'
        )
    
    def _import_grades_sheet(self, ws):
        """Import grades - read-only message"""
        QMessageBox.information(
            self, 'Import Not Available',
            'Grades are managed on the server.\n'
            'Use the web application to create/edit grades.'
        )
    
    def export_excel(self):
        """Export all data to Excel with multiple sheets"""
        from openpyxl import Workbook
        from datetime import datetime
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Export Data',
            str(config.EXCEL_DIR / f'frvv_export_{datetime.now():%Y%m%d_%H%M%S}.xlsx'),
            'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            total_records = 0
            
            # Athletes sheet
            athletes = self.db.get_all_athletes()
            ws_athletes = wb.create_sheet('Athletes')
            ws_athletes.append(['ID', 'First Name', 'Last Name', 'DOB', 'Mobile', 'Club', 'City', 'Grade', 'Status'])
            for athlete in athletes:
                ws_athletes.append([
                    athlete.get('id'),
                    athlete.get('first_name'),
                    athlete.get('last_name'),
                    athlete.get('date_of_birth'),
                    athlete.get('mobile_number'),
                    athlete.get('club_name'),
                    athlete.get('city_name'),
                    athlete.get('current_grade_name'),
                    athlete.get('status')
                ])
            total_records += len(athletes)
            
            # Competitions sheet
            competitions = self.db.get_all_competitions()
            ws_comps = wb.create_sheet('Competitions')
            ws_comps.append(['ID', 'Title', 'Start Date', 'End Date', 'City', 'Address', 'Type'])
            for comp in competitions:
                ws_comps.append([
                    comp.get('id'),
                    comp.get('title'),
                    comp.get('start_date'),
                    comp.get('end_date'),
                    comp.get('city_name'),
                    comp.get('address'),
                    comp.get('event_type')
                ])
            total_records += len(competitions)
            
            # Categories sheet
            categories = self.db.get_all_categories()
            ws_cats = wb.create_sheet('Categories')
            ws_cats.append(['ID', 'Competition', 'Name', 'Gender', 'Min Age', 'Max Age', 'Min Weight', 'Max Weight', 'Type'])
            for cat in categories:
                ws_cats.append([
                    cat.get('id'),
                    cat.get('competition_title'),
                    cat.get('name'),
                    cat.get('gender'),
                    cat.get('min_age'),
                    cat.get('max_age'),
                    cat.get('min_weight'),
                    cat.get('max_weight'),
                    cat.get('category_type')
                ])
            total_records += len(categories)
            
            # Matches sheet
            matches = self.db.get_all_matches()
            ws_matches = wb.create_sheet('Matches')
            ws_matches.append(['ID', 'Category', 'Round', 'Red Corner', 'Blue Corner', 'Winner', 'Status'])
            for match in matches:
                ws_matches.append([
                    match.get('id'),
                    match.get('category_name'),
                    match.get('round'),
                    match.get('athlete1_name'),
                    match.get('athlete2_name'),
                    match.get('winner_name'),
                    match.get('status')
                ])
            total_records += len(matches)
            
            # Clubs sheet
            clubs = self.db.get_all_clubs()
            ws_clubs = wb.create_sheet('Clubs')
            ws_clubs.append(['ID', 'Name', 'City', 'Address', 'Mobile', 'Website'])
            for club in clubs:
                ws_clubs.append([
                    club.get('id'),
                    club.get('name'),
                    club.get('city_name'),
                    club.get('address'),
                    club.get('mobile_number'),
                    club.get('website')
                ])
            total_records += len(clubs)
            
            # Grades sheet
            grades = self.db.get_all_grades()
            ws_grades = wb.create_sheet('Grades')
            ws_grades.append(['ID', 'Name', 'Rank Order', 'Type'])
            for grade in grades:
                ws_grades.append([
                    grade.get('id'),
                    grade.get('name'),
                    grade.get('rank_order'),
                    grade.get('grade_type')
                ])
            total_records += len(grades)
            
            wb.save(file_path)
            QMessageBox.information(
                self, 'Export Complete', 
                f'Exported {total_records} total records:\n'
                f'  - {len(athletes)} athletes\n'
                f'  - {len(competitions)} competitions\n'
                f'  - {len(categories)} categories\n'
                f'  - {len(matches)} matches\n'
                f'  - {len(clubs)} clubs\n'
                f'  - {len(grades)} grades'
            )
            
        except Exception as e:
            QMessageBox.critical(self, 'Export Failed', f'Error: {str(e)}')
    
    def show_about(self):
        """Show about dialog"""
        QMessageBox.about(
            self,
            'About',
            f'{config.APP_NAME}\n'
            f'Version {config.APP_VERSION}\n\n'
            f'Offline athlete management with sync to Django backend'
        )    
    def change_theme(self, theme_name: str):
        """Change application theme"""
        app = QApplication.instance()
        ThemeManager.apply_theme(app, theme_name)
        QMessageBox.information(
            self,
            'Theme Changed',
            f'{theme_name.capitalize()} mode activated!'
        )